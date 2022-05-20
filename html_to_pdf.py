# import pdfkit
#
# pdfkit.from_file('test.html', 'test.pdf')

import openpyxl

path = 'orders.xlsx'

work_obj = openpyxl.load_workbook(path)

sheet_obj= work_obj.active

# cell_obj = sheet_obj.cell(row=1, column=1)  #this will give first row and first column value
# print(cell_obj.value)

rows = sheet_obj.max_row
columns = sheet_obj.max_column


for i in range(1, rows + 1):
    cell_obj = sheet_obj.cell(row=i, column=3)
    print(cell_obj.value)

for i in range(1, rows+1):
    cell_obj = sheet_obj.cell(row=i, column=4)
    print(cell_obj.value)


print('The data of 5th column')
for i in range(1, columns+1):
    cell_obj = sheet_obj.cell(row=5, column=i)
    print(cell_obj.value)

cell_obj = sheet_obj['A1':'B10']
for cell1, cell2 in cell_obj:
    print(cell1.value, cell2.value)

wb = openpyxl.load_workbook('orders.xlsx')
sheet_obj = wb.active
c = sheet_obj['B4']
c.value = 'South'


sheet_obj['E11'] = '=SUM(E1:E10)'
wb.save('orders.xlsx')