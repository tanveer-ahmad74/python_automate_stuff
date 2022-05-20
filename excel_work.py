import openpyxl
from openpyxl.utils import  get_column_letter
#this is how u create a new xlsx file
# wb = openpyxl.Workbook()
#
# wb.save('New.xlsx')


path = 'New.xlsx'
work_obj = openpyxl.load_workbook(path)
sheet_obj = work_obj.active

for row in range(1, 10):
    for col in range(1, 6):
        char = get_column_letter(col)
        sheet_obj[char + str(row)] = char + str(row)

work_obj.save(path)